import csv
import io
import re
from datetime import UTC, date, datetime
from decimal import Decimal
from typing import Any

import httpx
from openpyxl import load_workbook

from app.core.config import settings


def fold(value: Any) -> str:
    text = str(value or "")
    return re.sub(r"\s+", " ", text.translate(str.maketrans("áéíóúüñÁÉÍÓÚÜÑ", "aeiouunAEIOUUN")).strip()).lower()


def decimal_value(value: Any) -> Decimal:
    if value is None or str(value).strip() in {"", "-", "..", "..."}:
        raise ValueError("Missing numeric value")
    text = str(value).strip().replace("€", "").replace("%", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    return Decimal(text)


def period_value(value: Any, frequency: str) -> date:
    if isinstance(value, datetime):
        return date(value.year, value.month if frequency != "annual" else 1, 1)
    if isinstance(value, date):
        return date(value.year, value.month if frequency != "annual" else 1, 1)
    text = fold(value)
    q = re.search(r"(\d{4}).*?([1-4])", text) if ("t" in text or "trimestre" in text) else None
    if q and frequency == "quarterly":
        return date(int(q.group(1)), (int(q.group(2)) - 1) * 3 + 1, 1)
    year = re.search(r"(19|20)\d{2}", text)
    if not year:
        raise ValueError(f"Cannot parse period {value}")
    return date(int(year.group(0)), 1, 1)


def geography(value: Any, level: str, code: Any | None = None) -> str:
    raw_code = re.sub(r"\D", "", str(code or ""))
    if level == "province" and raw_code:
        return f"PROV:{raw_code.zfill(2)}"
    if level == "ccaa" and raw_code:
        return f"CCAA:{raw_code.zfill(2)}"
    if level == "municipality" and raw_code:
        return f"MUN:{raw_code.zfill(5)}"
    if fold(value) in {"espana", "total nacional", "nacional"}:
        return "ES"
    return f"NAME:{fold(value)}"


async def download_rows(url: str, sheet_name: str | None = None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    async with httpx.AsyncClient(timeout=settings.http_timeout_seconds, follow_redirects=True) as client:
        response = await client.get(url)
        response.raise_for_status()
    content_type = response.headers.get("content-type", "")
    if url.lower().endswith((".xlsx", ".xlsm")) or "spreadsheet" in content_type:
        workbook = load_workbook(io.BytesIO(response.content), read_only=True, data_only=True)
        sheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
        values = list(sheet.iter_rows(values_only=True))
        header_index = next(i for i, row in enumerate(values) if sum(cell is not None for cell in row) >= 3)
        headers = [fold(cell) or f"column_{i}" for i, cell in enumerate(values[header_index])]
        rows = [dict(zip(headers, row)) for row in values[header_index + 1:] if any(cell is not None for cell in row)]
    else:
        text = response.content.decode(response.encoding or "utf-8-sig", errors="replace")
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=";,\t|")
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        rows = [{fold(k): v for k, v in row.items()} for row in reader]
    return rows, {"download_url":str(response.url),"content_type":content_type,"http_status":response.status_code,"retrieved_at":datetime.now(UTC).isoformat()}


def pick(row: dict[str, Any], aliases: list[str]) -> Any:
    normalized = {fold(k): v for k, v in row.items()}
    for alias in aliases:
        target = fold(alias)
        for key, value in normalized.items():
            if target == key or target in key:
                return value
    return None
